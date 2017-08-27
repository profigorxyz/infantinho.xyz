# TODO construir uma loading bar para quando está a carregar a tabela
# e ou a construir os ficheiros para download
import os
from django.conf import settings
from django.shortcuts import render  # , redirect
from .models import PresRec
from django.contrib import messages
from herd.models import Teacher, Subject, Student, Grade
from django.core.files.storage import FileSystemStorage, DefaultStorage, Storage
from openpyxl import workbook, load_workbook
from openpyxl.drawing.image import Image
# from django.http import Http404
# from django.core.files.storage import default_storage
from .forms import PresForm  # DateForm, PrintDateForm
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
# from django.core import serializers
import datetime


@login_required(login_url='/login/google-oauth2/?next=/')
def pres_create(request):
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request,
                       'Tem de ter autorização para poder marcar presenças.')
        return render(request, 'blank.html')
    form = PresForm()
    subject = Subject.objects.filter(teacher__user=request.user.id)
    context = {
        'form': form,
        'subject': subject,
    }
    if request.method == 'POST':
        students = Student.objects.filter(grade__exact=request.POST['grade'])
        date = request.POST['date_submit'].split('\'')[0]
        subject = Subject.objects.get(id=request.POST['subject'])
        for student in students:
            try:
                if request.POST[str(student.id)]:
                    absent = 1
            except KeyError:
                absent = 0
            p = PresRec(
                student=student,
                is_absent=absent,
                subject=subject,
                date=date,
            )
            p.save()
        messages.success(request,
                         'Presenças de {} marcadas para o dia {}.'.format(
                             subject,
                             date))
    return render(request, 'record/create_pres.html', context)


@login_required(login_url='/login/google-oauth2/?next=/')
def pres_read(request):
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request, 'Tem de ter\
            autorização para poder verificar presenças.')
        return render(request, 'blank.html')
    subject = Subject.objects.filter(teacher__user=request.user.id)
    context = {
        'subject': subject,
    }
    return render(request, 'record/read_pres.html', context)


@login_required(login_url='/login/google-oauth2/?next=/')
def pres_print(request):
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request, 'Tem de ter\
            autorização para poder verificar presenças.')
        return render(request, 'blank.html')
    subject = Subject.objects.filter(teacher__user=request.user.id)
    context = {
        'subject': subject,
    }
    return render(request, 'record/print_pres.html', context)


@login_required(login_url='/login/google-oauth2/?next=/')
def skill_create(request):
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request, 'Tem de ter\
            autorização para poder verificar presenças.')
        return render(request, 'blank.html')
    subject = Subject.objects.filter(teacher__user=request.user.id)
    context = {
        'subject': subject,
    }
    return render(request, 'record/read_pres.html', context)


@login_required(login_url='/login/google-oauth2/?next=/')
def skill_read(request):
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request, 'Tem de ter\
            autorização para poder verificar presenças.')
        return render(request, 'blank.html')
    subject = Subject.objects.filter(teacher__user=request.user.id)
    context = {
        'subject': subject,
    }
    return render(request, 'record/read_skill.html', context)


@login_required(login_url='/login/google-oauth2/?next=/')
def skill_update(request):
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request, 'Tem de ter\
            autorização para poder verificar presenças.')
        return render(request, 'blank.html')
    subject = Subject.objects.filter(teacher__user=request.user.id)
    context = {
        'subject': subject,
    }
    return render(request, 'record/update_skill.html', context)


@login_required(login_url='/login/google-oauth2/?next=/')
def skill_delete(request):
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request, 'Tem de ter\
            autorização para poder verificar presenças.')
        return render(request, 'blank.html')
    subject = Subject.objects.filter(teacher__user=request.user.id)
    context = {
        'subject': subject,
    }
    return render(request, 'record/delete_skill.html', context)


@login_required(login_url='/login/google-oauth2/?next=/')
def get_print_url(request):
    subject = int(request.GET.get('subject', None))
    if not subject:
        messages.error(request,
                       'Tem de escolher uma turma.')
        return render(request, 'blank.html')
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request,
                       'Tem de ter autorização para aceder.')
        return render(request, 'blank.html')
    storage = DefaultStorage()
    flpres = storage.open('presence.xlsx', mode='rb')
    wb = load_workbook(flpres)
    ws = wb['Folha1']
    now = datetime.datetime.now()
    if now.month < 7:
        end_date = datetime.date(now.year, 7, 1)
        start_date = datetime.date(now.year - 1, 9, 1)
    else:
        end_date = datetime.date(now.year + 1, 7, 1)
        start_date = datetime.date(now.year, 9, 1)
    p = PresRec.objects.filter(
        subject__exact=subject,
        date__range=(start_date, end_date)
    ).order_by('-date')
    subject = Subject.objects.get(pk__exact=subject)
    grade = Grade.objects.get(pk__exact=subject.grade.pk)
    students = Student.objects.filter(
        grade__exact=grade).order_by('user__first_name', 'user__last_name')
    for student in students:
        ws['A' + get_y(student.list_number - 1, 9)] = '{} {}'.format(
            student.user.first_name, student.user.last_name)
    for p in p:
        if p.is_absent == 0:
            pres = 'P'
        elif p.is_absent == 1:
            pres = 'F'
        ws[get_x(p.date.day) + get_y(
            p.student.list_number - 1,
            p.date.month)] = pres
    ws = wb['Folha1']
    fllogo = storage.open('logotipo.png', mode='rb')
    img = Image(fllogo)
    ws.add_image(img, 'L3')
    ws['A35'] = '{}'.format(subject)
    ws['A62'] = '{} {} / {}'.format(
        'Ano letivo de ', start_date.year, end_date.year)
    filename = ''.join('{}.xlsx'.format(subject).split())
    fs = FileSystemStorage()
    fs = fs.open(filename, mode='wb+')
    # This dont work properly
    # fs = storage.open('print/{}'.format(filename), mode='wb+')
    wb.save(fs)
    storage.save(filename, fs)
    fs.close()
    fllogo.close()
    flpres.close()
    os.remove(os.path.join(settings.MEDIA_ROOT, filename))
    file_url = storage.url(filename)
    data = []
    data.append({
        'url': file_url,
        'name': '{}'.format(filename).upper(),
    })
    return JsonResponse(data, safe=False)


@login_required(login_url='/login/google-oauth2/?next=/')
def get_students(request):
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request,
                       'Tem de ter autorização para aceder.')
        return render(request, 'blank.html')
    subject = request.GET.get('subject', None)
    subject = Subject.objects.get(pk__exact=subject)
    grade = Grade.objects.get(pk__exact=subject.grade.pk)
    students = Student.objects.filter(
        grade__exact=grade).order_by('user__first_name', 'user__last_name')
    subject = str(subject.id)
    studic = []
    for student in students:
        studic.append({
                      'id': student.id,
                      'grade': student.grade_id,
                      'name': '{}'.format(student).upper(),
                      'jpg': student.photo.face.url,
                      'subject': subject
                      }
                      )
    data = studic
    return JsonResponse(data, safe=False)


@login_required(login_url='/login/google-oauth2/?next=/')
def get_presents(request):
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request,
                       'Tem de ter autorização para aceder.')
        return render(request, 'blank.html')
    subject = request.GET.get('subject', None)
    subject = Subject.objects.get(pk__exact=subject)
    grade = Grade.objects.get(pk__exact=subject.grade.pk)
    students = Student.objects.filter(
        grade__exact=grade).order_by('user__first_name', 'user__last_name')
    subjtitle = str(subject)
    now = datetime.datetime.now()
    if now.month < 12 and now.month > 8:
        start_date = datetime.date(now.year, 9, 1)
        end_date = datetime.date(now.year, now.month, now.day)
    elif now.month < 3:
        start_date = datetime.date(now.year - 1, now.month - 3, 1)
        end_date = datetime.date(now.year, now.month, now.day)
    elif now.month > 6:
        start_date = datetime.date(now.year, 4, 1)
        end_date = datetime.date(now.year, 6, 30)
    else:
        start_date = datetime.date(now.year, now.month - 3, 1)
        end_date = datetime.date(now.year, now.month, now.day)
    studic = []
    for student in students:
        pres = PresRec.objects.filter(
            subject__exact=subject,
            student__user__exact=student.user,
            date__range=[
                start_date,
                end_date
            ],
            is_absent=0).count()
        absent = PresRec.objects.filter(
            subject__exact=subject,
            student__user__exact=student.user,
            date__range=[
                start_date,
                end_date
            ],
            is_absent=1).count()
        studic.append({
                      'name': '{}'.format(student),
                      'pres': pres,
                      'absent': absent,
                      'start_date': start_date,
                      'end_date': end_date,
                      'subject': subjtitle
                      }
                      )
    data = studic
    return JsonResponse(data, safe=False)


def get_x(day):
    switcher = {
        1: "C",
        2: "D",
        3: "E",
        4: "F",
        5: "G",
        6: "H",
        7: "I",
        8: "J",
        9: "K",
        10: "L",
        11: "M",
        12: "N",
        13: "O",
        14: "P",
        15: "Q",
        16: "R",
        17: "S",
        18: "T",
        19: "U",
        20: "V",
        21: "W",
        22: "X",
        23: "Y",
        24: "Z",
        25: "AA",
        26: "AB",
        27: "AC",
        28: "AD",
        29: "AE",
        30: "AF",
        31: "AG"
    }
    return switcher.get(day)


def get_y(counter, month):
    y = 90 + (counter * 10)
    if month > 8 and month < 13:
        y = y + counter + (month - 7)
    else:
        y = y + counter + month + 5
    return str(y)

    # if now.month < 3:
    #     end_month = datetime.date(now.year, 7, 1)
    #     start_date = datetime.date(now.year - 1, 9, 1)
    # else:
    #     end_date = datetime.date(now.year + 1, 7, 1)
    #     start_date = datetime.date(now.year, 9, 1)


#     if request.method == 'POST' and request.POST['stage'] == '1':
#         import datetime
#         from openpyxl import workbook, load_workbook
#         from openpyxl.drawing.image import Image
# #        from openpyxl.drawing.image import Image // implementar logo
#         flpres = open(os.path.join(LOCAL_FILES, 'presence.xlsx'), 'rb')
#         wb = load_workbook(flpres)
#         ws = wb['PRES']
#         now = datetime.datetime.now()
#         if now.month < 7:
#             end_date = datetime.date(now.year, 7, 1)
#             start_date = datetime.date(now.year - 1, 9, 1)
#         else:
#             end_date = datetime.date(now.year + 1, 7, 1)
#             start_date = datetime.date(now.year, 9, 1)
#         p = Presence.objects.filter(
#                                    subject__exact=int(request.POST['subject']),
#                                    date__range=(start_date, end_date)
#                                    ).order_by('-date')

#         def gotx(day):
#             switcher = {1: "C", 2: "D", 3: "E", 4: "F", 5: "G", 6: "H", 7: "I", 8: "J", 9: "K", 10: "L", 11: "M", 12: "N",
#                                 13: "O", 14: "P", 15: "Q", 16: "R", 17: "S", 18: "T", 19: "U", 20: "V", 21: "W", 22: "X", 23: "Y", 24: "Z",
#                                 25: "AA", 26: "AB", 27: "AC", 28: "AD", 29: "AE", 30: "AF", 31: "AG"}
#             return switcher.get(day)

#         def goty(counter,month):
#             y = 56 + (counter * 10)
#             if month > 8 and month < 13:
#                 y = y + counter + (month - 7)
#             else:
#                 y = y + counter + month + 5
#             return str(y)
#         gsub = Subject.objects.get(id=request.POST['subject'])
#         students = Student.objects.filter(grade__exact=gsub.grade)
#         for student in students:
#             ws['A' + goty(student.list_number-1, 9)] = '{} {}'.format(student.name.first_name, student.name.last_name)
#         for p in p:
#             if p.is_absent == 0:
#                 pres = 'P'
#             elif p.is_absent == 1:
#                 pres = 'F'
#             ws[gotx(p.date.day) + goty(
#                                          p.student.list_number-1,
#                                          p.date.month)] = pres
#         ws = wb['PRES']
#         fllogo = open(os.path.join(LOCAL_FILES, 'logotipo.png'), 'rb')
#         img = Image(fllogo)
#         ws.add_image(img, 'L3')
#         ws['A22'] = '{}'.format(gsub.name)
#         ws['A27'] = '{}'.format(gsub.grade.name)
#         ws['A44']  = '{} {} / {}'.format('Ano letivo de ', start_date.year, end_date.year)
#         wb.save('{}/{}_{}.xlsx'.format(LOCAL_FILES, gsub.name, gsub.grade.name))
#         subject = Subject.objects.all()
#         return render(request, 'record/report.html', {
#                   'errors': errors,
#                   'subject': subject,
#                   })
#     else:
#         subject = Subject.objects.all()
#         return render(request, 'record/report.html', {
#                   'errors': errors,
#                   'subject': subject,
#                   })


# @login_required(login_url='/login/google-oauth2/?next=/')
# def registry(request):
#     try:
#         teacher = Teacher.objects.filter(name__exact=request.user.id).first()
#     except:
#         teacher = None
#     if teacher:
#         context = {
#             'teacher': teacher,
#         }
#         return render(request, 'record/entrypoint.html', context)
#     else:
#         context = {
#         }
#         return render(request, 'record/entrypoint.html', context)
#     return render(request, 'blank.html')


# @login_required(login_url='/login/google-oauth2/?next=/')
# def print_report(request):
#     teacher = Teacher.objects.filter(name__exact=request.user.id).first()
#     if not teacher:
#         messages.error(request, 'Tem de ter autorização para poder verificar as presenças.')
#         return render(request, 'blank.html')
#     subject = Subject.objects.filter(teacher__name=request.user.id)
#     form = PrintDateForm(request.POST or None)
#     if form.is_valid():
#         grade = int(request.POST['subject'].split("-")[1])
#         subject = int(request.POST['subject'].split("-")[0])
#         students = Student.objects.filter(grade__exact=grade)
#         teacher = Teacher.objects.filter(name__exact=request.user.id, subject__exact=int(request.POST['subject'].split("-")[0])).first()
#         printout = {}
#         for student in students:
#             pres = Presence.objects.filter(
#                 subject__exact=subject,
#                 student__name__exact=student.name,
#                 date__range=[
#                     request.POST['initialdate'],
#                     request.POST['finaldate']
#                 ],
#                 is_absent=0
#             ).count()
#             absent = Presence.objects.filter(
#                 subject__exact=subject,
#                 student__name__exact=student.name,
#                 date__range=[
#                     request.POST['initialdate'],
#                     request.POST['finaldate']
#                 ],
#                 is_absent=1
#             ).count()
#             printout[student.list_number] = {
#                 'name': '{} {}'.format(
#                     student.name.first_name,
#                     student.name.last_name
#                 ),
#                 'presents': pres,
#                 'absents': absent,
#             }
#         context = {
#             'teacher': teacher,
#             'initialdate': request.POST['initialdate'],
#             'finaldate': request.POST['finaldate'],
#             'subject': subject,
#             'print': sorted(printout.items()),
#         }
#         return render(request, 'record/print.html', context)
#     context = {
#         'subject': subject,
#         'form': form,
#     }
#     return render(request, 'record/printform.html', context)
