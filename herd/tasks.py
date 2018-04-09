from __future__ import absolute_import, unicode_literals
from celery import task, shared_task
from django.contrib.auth.models import User
from .models import Student, Grade, Subject
from openpyxl import load_workbook
import zipfile
import os
import datetime
import json
import glob
import shutil
import calendar
from django.core.files.images import ImageFile
from django.core.files.storage import FileSystemStorage
from django.core.files import File
from django.conf import settings
from django.core.mail import EmailMessage
from record.models import PresRec


@shared_task
def send_report():
    email = EmailMessage(
        'Relatório de Presenças',
        'Segue em anexo o relatório de presenças deste mês.',
        'infantinho@colegioinfante.info',
        [
            'prof.igor@colegioinfante.info',
            'jorge@colegioinfante.info',
            'secretariaph2012@gmail.com',
        ],
    )
    teacheruser = 395
    subject = Subject.objects.filter(teacher__user=teacheruser)
    grade = list()
    for s in subject:
        [grade.append(pk.get('pk')) for pk in s.grade.values('pk')]
    students = Student.objects.filter(
        grade__in=grade).order_by('number')
    now = datetime.datetime.now()
    clear_rec_db(now.year, now.month, now.day)
    mdays = calendar.monthrange(now.year, now.month)[1]
    weekday = calendar.weekday(now.year, now.month, now.day)
    if mdays != now.day and weekday != 4:
        return 'not lastday, and not friday day: {}'.format(now.day)
    if (mdays - now.day) > 2:
        return 'not last friday of the month, day: {}'.format(now.day)
    if now.hour != 10:
        return 'last friday but time is: {}:{}'.format(now.hour, now.minute)
    if now.minute <= 10 or now.minute > 20:
        return 'last friday but time is: {}:{}'.format(now.hour, now.minute)
    start_date = datetime.date(now.year, now.month, 1)
    end_date = datetime.date(now.year, now.month, mdays)
    studic = []
    for student in students:
        pres = PresRec.objects.filter(
            subject=Subject.objects.get(
                name='secretaria',
                grade=student.grade
            ),
            student__user__exact=student.user,
            date__range=[
                start_date,
                end_date
            ],
            is_absent=0).count()
        absent = PresRec.objects.filter(
            subject__exact=Subject.objects.get(
                name='secretaria',
                grade=student.grade
            ),
            student__user__exact=student.user,
            date__range=[
                start_date,
                end_date
            ],
            is_absent=1).count()
        studic.append({
                      'name': '{}'.format(student),
                      'number': student.number,
                      'pres': pres,
                      'absent': absent,
                      'start_date': str(start_date),
                      'end_date': str(end_date),
                      }
                      )
    jsoncontent = json.dumps(studic)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'pres'
    row_num = 0
    columns = [
        (u'Número', 6, 'A'),
        (u'Nome', 25, 'B'),
        #(u'Presenças', 7, 'C'),
        (u'Faltas', 7, 'C')
    ]
    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.font = openpyxl.styles.fonts.Font(bold=True)
        ws.column_dimensions[columns[col_num][2]].width = columns[col_num][1]
    for e in studic:
        row_num += 1
        row = [
            e.get('number'),
            e.get('name'),
            #e.get('pres'),
            e.get('absent'),
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
    fs = FileSystemStorage()
    f = fs.open('pres.xlsx', mode='wb+')
    f = File(f)
    wb.save(f)
    f.close
    email.attach('pres.json', jsoncontent, 'application/json')
    email.attach_file(os.path.join(settings.MEDIA_ROOT, 'pres.xlsx'))
    email.send()
    return 'task done at {}:{}'.format(now.hour, now.minute)


def clear_rec_db(year, month, days):
    teacheruser = 395
    month = month
    year = year
    days = days
    start_date = datetime.date(year, month, 1)
    end_date = datetime.date(year, month, days)
    subject = Subject.objects.filter(teacher__user=teacheruser)
    grade = list()
    for s in subject:
        [grade.append(pk.get('pk')) for pk in s.grade.values('pk')]
    students = Student.objects.filter(
        grade__in=grade).order_by('number')
    for student in students:
        for edate in PresRec.objects.filter(
            student__user__exact=student.user,
            subject=Subject.objects.get(
                name='secretaria',
                grade=student.grade
            ),
            date__range=[
                start_date,
                end_date
            ]
        ).values_list('date', flat=True).distinct():
            plist = PresRec.objects.filter(
                student__user__exact=student.user,
                subject=Subject.objects.get(
                    name='secretaria',
                    grade=student.grade
                ),
                date__exact=edate).values_list('id', flat=True)
            plist = list(plist)[1:]
            PresRec.objects.filter(pk__in=plist).delete()
    return 0


@shared_task
def build_school(x):
    wb = load_workbook(
        filename=x,
        read_only=True
    )
    for book in wb:
        ngrade = '{}'.format(book.title).upper()
        b = 1
        grades = {}
        grades[ngrade] = {}
        for row in book.rows:
            grades[ngrade]['{}'.format(row[0].value).rstrip('.')] = {
                'number': '{}'.format(row[0].value).rstrip('.'),
                'list_number': b,
                'name': row[1].value
            }
            b += 1
        [pop_student.delay(grades[ngrade][v], ngrade) for v in grades[ngrade]]
        for s in grades[ngrade]:
            try:
                student_number = int(grades[ngrade][s].get('number'))
            except ValueError:
                continue
            student = Student.objects.filter(
                number__exact=student_number).first()
            if '{}'.format(str(student_number)) not in grades[ngrade]:
                student.grade = Grade.objects.filter(name__exact='EX').first()
                student.save()
            print('Aluno {} criado ou atualizado.'.format(s))
    files = glob.glob('{}0*.*'.format(settings.MEDIA_ROOT))
    for f in files:
        os.remove(f)


@shared_task
def pop_student(x, y):
    student = x
    grade = y
    if not Grade.objects.filter(name__exact=grade).exists():
        gobj = Grade(name=grade)
        gobj.save()
    else:
        gobj = Grade.objects.get(name=grade)
    fs = FileSystemStorage()
    st_name = student.get('name')
    if not st_name:
        return 0
    st_name = st_name.rstrip(' ')
    f_name = st_name.split(' ', 1)[0]
    namelist = st_name.split(' ', -1)
    l_name = namelist[len(namelist) - 1]
    st_number = student.get('number')
    username = '{}_{}'.format(f_name, student.get('number'))
    if not User.objects.filter(username__exact=username).exists():
        u = User.objects.create_user(
            username,
            st_number + '@colegioinfante.info',
            'aluno' + st_number,
        )
        u.first_name = f_name
        u.last_name = l_name
        u.save()
    else:
        u = User.objects.get(username=username)
    fname = '{}.jpg'.format(st_number)
    try:
        fimg = fs.open(fname, mode='rb')
    except FileNotFoundError:
        default = '{}'.format(os.path.join(settings.MEDIA_ROOT,
                                           'default.jpg'))
        shutil.copy(default, os.path.join(settings.MEDIA_ROOT, fname))
        fimg = fs.open(fname, mode='rb')
    f = ImageFile(fimg)
    if not Student.objects.filter(number__exact=int(st_number)).exists():
        s = Student(
            user=u,
            number=int(st_number),
            list_number=student.get('list_number'),
            grade=gobj,
            photo=f,
        )
        s.save()
    else:
        s = Student.objects.get(number=int(st_number))
        s.grade = gobj
        s.photo = f
        s.save()
    fimg.close()
    f.close()
    files = glob.glob('{}{}*.jpg'.format(settings.MEDIA_ROOT, fname))
    for f in files:
        os.remove(f)


@task(name="create_students")
def create_students(fzip, xlsx):
    fjpg = zipfile.ZipFile(fzip, 'r')
    fjpg.extractall(settings.MEDIA_ROOT)
    build_school.delay(xlsx)
