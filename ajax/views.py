import os
import datetime
from text_unidecode import unidecode
from django.http import JsonResponse
from django.conf import settings
from django.shortcuts import render
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage, DefaultStorage
from django.core.files.base import ContentFile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from herd.models import Teacher, Subject, Student, Grade
from record.models import PresRec


def get_x(day):
    switcher = {
        1: "C",
        2: "D",
        3: "E",
        4: "F",
        5: "G",
        6: "H",
        7: "I",
        8: "J",
        9: "K",
        10: "L",
        11: "M",
        12: "N",
        13: "O",
        14: "P",
        15: "Q",
        16: "R",
        17: "S",
        18: "T",
        19: "U",
        20: "V",
        21: "W",
        22: "X",
        23: "Y",
        24: "Z",
        25: "AA",
        26: "AB",
        27: "AC",
        28: "AD",
        29: "AE",
        30: "AF",
        31: "AG"
    }
    return switcher.get(day)


def get_y(counter, month):
    y = 90 + (counter * 10)
    if month > 8 and month < 13:
        y = y + counter + (month - 7)
    else:
        y = y + counter + month + 5
    return str(y)


@login_required(login_url='/login/google-oauth2/?next=/')
def get_print_url(request):
    subject = int(request.GET.get('subject', None))
    if not subject:
        messages.error(request,
                       'Tem de escolher uma turma.')
        return render(request, 'blank.html')
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request,
                       'Tem de ter autorização para aceder.')
        return render(request, 'blank.html')
    storage = DefaultStorage()
    flpres = storage.open('presence.xlsx', mode='rb')
    wb = load_workbook(flpres)
    ws = wb['Folha1']
    now = datetime.datetime.now()
    if now.month < 7:
        end_date = datetime.date(now.year, 7, 1)
        start_date = datetime.date(now.year - 1, 9, 1)
    else:
        end_date = datetime.date(now.year + 1, 7, 1)
        start_date = datetime.date(now.year, 9, 1)
    p = PresRec.objects.filter(
        subject__exact=subject,
        date__range=(start_date, end_date)
    ).order_by('-date')
    subject = Subject.objects.get(pk__exact=subject)
    grade = Grade.objects.get(pk__exact=subject.grade.pk)
    students = Student.objects.filter(
        grade__exact=grade).order_by('user__first_name', 'user__last_name')
    for student in students:
        ws['A' + get_y(student.list_number - 1, 9)] = '{} {}'.format(
            student.user.first_name, student.user.last_name)
    for p in p:
        if p.is_absent == 0:
            pres = 'P'
        elif p.is_absent == 1:
            pres = 'F'
        ws[get_x(p.date.day) + get_y(
            p.student.list_number - 1,
            p.date.month)] = pres
    ws = wb['Folha1']
    fllogo = storage.open('logotipo.png', mode='rb')
    img = Image(fllogo)
    ws.add_image(img, 'L3')
    ws['A35'] = '{}'.format(subject)
    ws['A62'] = '{} {} / {}'.format(
        'Ano letivo de ', start_date.year, end_date.year)
    filename = ''.join('{}.xlsx'.format(subject).split())
    fs = FileSystemStorage()
    fs = fs.open(filename, mode='wb+')
    # This dont work properly
    # fs = storage.open('print/{}'.format(filename), mode='wb+')
    # fs = ContentFile()
    wb.save(fs)
    storage.save(filename, fs)
    fs.close()
    fllogo.close()
    flpres.close()
    os.remove(os.path.join(settings.MEDIA_ROOT, filename))
    file_url = storage.url(filename)
    data = []
    data.append({
        'url': file_url,
        'name': '{}'.format(filename).upper(),
    })
    return JsonResponse(data, safe=False)


@login_required(login_url='/login/google-oauth2/?next=/')
def get_students(request):
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request,
                       'Tem de ter autorização para aceder.')
        return render(request, 'blank.html')
    grade = request.GET.get('grade', None)
    subject = request.GET.get('subject', None)
    if not grade:
        sub = Subject.objects.filter(pk=subject).first()
        grade = [
            gpk.get('pk')
            for gpk in sub.grade.values('pk')
        ]
    if sub.club is True:
        sts = Student.objects.filter(club=sub.id)
    else:
        sts = Student.objects.filter(
            grade__in=grade).order_by(
                'grade',
                'user__first_name',
                'user__last_name'
        )
    allgd = Grade.objects.all()
    grades = {}
    students = {}
    team = {}
    for agd in allgd:
        grades[agd.name] = {
            'id': agd.id,
            'name': agd.name,
        }
    first = True
    for student in sts:
        if first:
            first = False
            gname = student.grade.name
            team[gname] = {}
            team[gname]['grade'] = student.grade_id
            team[gname]['subject'] = subject
            team[gname]['name'] = student.grade.name
        else:
            ngname = student.grade.name
            if ngname != gname:
                gname = student.grade.name
                team[gname] = {}
                team[gname]['grade'] = student.grade_id
                team[gname]['subject'] = subject
                team[gname]['name'] = student.grade.name
        name = ''.join('{}'.format(student).split())
        name = unidecode(name)
        students['{}'.format(name)] = {
            'id': student.id,
            'name': '{}'.format(student).upper(),
            'jpg': student.photo.face.url,
            'team': student.grade.name
        }
    data = {
        'grade': grades,
        'student': students,
        'class': team,
    }
    jsondump = {
        'sort_keys': True,
    }
    return JsonResponse(data, json_dumps_params=jsondump)
    # grades = []
    # students = []
    # for agd in allgd:
    #     grades.append({
    #         agd.id: {
    #             'id': agd.id,
    #             'name': agd.name
    #         }
    #     })
    # for student in sts:
    #     students.append({
    #         student.id: {
    #             'id': student.id,
    #             'grade': student.grade_id,
    #             'name': '{}'.format(student).upper(),
    #             'jpg': student.photo.face.url,
    #             'subject': subject
    #         }
    #     })

@login_required(login_url='/login/google-oauth2/?next=/')
def get_presents(request):
    teacher = Teacher.objects.filter(user__exact=request.user.id).first()
    if not teacher:
        messages.error(request,
                       'Tem de ter autorização para aceder.')
        return render(request, 'blank.html')
    subject = request.GET.get('subject', None)
    subject = Subject.objects.get(pk__exact=subject)
    grade = Grade.objects.get(pk__exact=subject.grade.pk)
    students = Student.objects.filter(
        grade__exact=grade).order_by('user__first_name', 'user__last_name')
    subjtitle = str(subject)
    now = datetime.datetime.now()
    if now.month < 12 and now.month > 8:
        start_date = datetime.date(now.year, 9, 1)
        end_date = datetime.date(now.year, now.month, now.day)
    elif now.month < 3:
        start_date = datetime.date(now.year - 1, now.month - 3, 1)
        end_date = datetime.date(now.year, now.month, now.day)
    elif now.month > 6:
        start_date = datetime.date(now.year, 4, 1)
        end_date = datetime.date(now.year, 6, 30)
    else:
        start_date = datetime.date(now.year, now.month - 3, 1)
        end_date = datetime.date(now.year, now.month, now.day)
    studic = []
    for student in students:
        pres = PresRec.objects.filter(
            subject__exact=subject,
            student__user__exact=student.user,
            date__range=[
                start_date,
                end_date
            ],
            is_absent=0).count()
        absent = PresRec.objects.filter(
            subject__exact=subject,
            student__user__exact=student.user,
            date__range=[
                start_date,
                end_date
            ],
            is_absent=1).count()
        studic.append({
                      'name': '{}'.format(student),
                      'pres': pres,
                      'absent': absent,
                      'start_date': start_date,
                      'end_date': end_date,
                      'subject': subjtitle
                      }
                      )
    data = studic
    return JsonResponse(data, safe=False)
